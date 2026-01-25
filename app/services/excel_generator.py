import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.schemas.cases import Case


def generate_excel_from_mock_data(case: Case, mock_data: dict) -> bytes:
    """
    Generate an Excel workbook from mock data.
    Returns bytes of the Excel file ready for download.
    """
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    if "payroll_runs" in mock_data:
        _create_payroll_sheets(wb, case, mock_data)
    elif "mrr_forecast" in mock_data:
        _create_revenue_forecast_sheets(wb, case, mock_data)
    elif "expenses" in mock_data:
        _create_expense_audit_sheets(wb, case, mock_data)
    elif "transactions" in mock_data:
        _create_financial_sheets(wb, case, mock_data)
    else:
        _create_analysis_sheets(wb, case, mock_data)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _create_financial_sheets(wb, case: Case, mock_data: dict):
    """Create sheets for financial data."""
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "Financial Case Summary"
    ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_summary.merge_cells('A1:B1')
    
    ws_summary['A3'] = "Company"
    ws_summary['B3'] = mock_data['company']
    ws_summary['A4'] = "Case Title"
    ws_summary['B4'] = case.title
    ws_summary['A5'] = "Report Date"
    ws_summary['B5'] = mock_data['report_date']
    ws_summary['A6'] = "Generated On"
    ws_summary['B6'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    ws_summary['A8'] = "Summary Metrics"
    ws_summary['A8'].font = Font(bold=True, size=12)
    
    ws_summary['A9'] = "Total Revenue"
    ws_summary['B9'] = mock_data['summary']['total_revenue']
    ws_summary['A10'] = "Total Expenses"
    ws_summary['B10'] = mock_data['summary']['total_expense']
    ws_summary['A11'] = "AR Total Outstanding"
    ws_summary['B11'] = mock_data['summary']['ar_total']
    ws_summary['A12'] = "AR Overdue"
    ws_summary['B12'] = mock_data['summary']['ar_overdue']
    ws_summary['A13'] = "Projected Balance"
    ws_summary['B13'] = mock_data['summary']['projected_balance']
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # Transactions sheet
    ws_trans = wb.create_sheet("Transactions", 1)
    _write_data_table(ws_trans, mock_data['transactions'], 
                      ["date", "description", "category", "amount", "status"],
                      ["Date", "Description", "Category", "Amount", "Status"])
    
    # AR Aging sheet
    ws_ar = wb.create_sheet("AR Aging", 2)
    _write_data_table(ws_ar, mock_data['ar_aging'],
                      ["customer", "invoice_number", "invoice_date", "due_date", "amount", "days_overdue", "status"],
                      ["Customer", "Invoice #", "Invoice Date", "Due Date", "Amount", "Days Overdue", "Status"])
    
    # Cash Flow sheet
    ws_cf = wb.create_sheet("Cash Flow", 3)
    _write_data_table(ws_cf, mock_data['cash_flow_forecast'],
                      ["week", "inflows", "outflows", "net", "balance"],
                      ["Week", "Inflows", "Outflows", "Net", "Balance"])


def _create_expense_audit_sheets(wb, case: Case, mock_data: dict):
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "Expense Audit Summary"
    ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_summary.merge_cells('A1:B1')

    ws_summary['A3'] = "Company"; ws_summary['B3'] = mock_data['company']
    ws_summary['A4'] = "Case Title"; ws_summary['B4'] = case.title
    ws_summary['A5'] = "Report Date"; ws_summary['B5'] = mock_data['report_date']
    ws_summary['A8'] = "Total Expense"; ws_summary['B8'] = mock_data['summary']['total_expense']
    ws_summary['A9'] = "Over Budget Items"; ws_summary['B9'] = mock_data['summary']['over_budget_items']
    ws_summary['A10'] = "Under Budget Items"; ws_summary['B10'] = mock_data['summary']['under_budget_items']
    ws_summary.column_dimensions['A'].width = 25; ws_summary.column_dimensions['B'].width = 22

    ws_exp = wb.create_sheet("Expenses", 1)
    _write_data_table(ws_exp, mock_data['expenses'],
                      ["date", "vendor", "amount", "category", "department", "budget", "variance"],
                      ["Date", "Vendor", "Amount", "Category", "Department", "Budget", "Variance"])

    ws_var = wb.create_sheet("Variance by Dept", 2)
    _write_data_table(ws_var, mock_data['variance_summary'],
                      ["department", "variance"],
                      ["Department", "Variance"])


def _create_revenue_forecast_sheets(wb, case: Case, mock_data: dict):
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "Revenue Forecast Summary"
    ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_summary.merge_cells('A1:B1')

    ws_summary['A3'] = "Company"; ws_summary['B3'] = mock_data['company']
    ws_summary['A4'] = "Case Title"; ws_summary['B4'] = case.title
    ws_summary['A5'] = "Report Date"; ws_summary['B5'] = mock_data['report_date']
    ws_summary['A8'] = "Ending MRR"; ws_summary['B8'] = mock_data['summary']['ending_mrr']
    ws_summary['A9'] = "Avg New MRR"; ws_summary['B9'] = mock_data['summary']['avg_new_mrr']
    ws_summary['A10'] = "Avg Churn MRR"; ws_summary['B10'] = mock_data['summary']['avg_churn_mrr']
    ws_summary.column_dimensions['A'].width = 24; ws_summary.column_dimensions['B'].width = 22

    ws_mrr = wb.create_sheet("MRR Forecast", 1)
    _write_data_table(ws_mrr, mock_data['mrr_forecast'],
                      ["month", "new_mrr", "churn_mrr", "expansion_mrr", "total_mrr"],
                      ["Month", "New MRR", "Churn MRR", "Expansion MRR", "Total MRR"])

    ws_scen = wb.create_sheet("Scenarios", 2)
    _write_data_table(ws_scen, mock_data['scenarios'],
                      ["scenario", "growth_rate", "churn_change", "projected_mrr"],
                      ["Scenario", "Growth Rate", "Churn Change", "Projected MRR"])


def _create_payroll_sheets(wb, case: Case, mock_data: dict):
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "Payroll Reconciliation"
    ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_summary.merge_cells('A1:B1')

    ws_summary['A3'] = "Company"; ws_summary['B3'] = mock_data['company']
    ws_summary['A4'] = "Case Title"; ws_summary['B4'] = case.title
    ws_summary['A5'] = "Report Date"; ws_summary['B5'] = mock_data['report_date']
    ws_summary['A8'] = "Total Gross"; ws_summary['B8'] = mock_data['summary']['total_gross']
    ws_summary['A9'] = "Total Taxes"; ws_summary['B9'] = mock_data['summary']['total_taxes']
    ws_summary['A10'] = "Total Benefits"; ws_summary['B10'] = mock_data['summary']['total_benefits']
    ws_summary['A11'] = "Employer Taxes"; ws_summary['B11'] = mock_data['summary']['employer_taxes']
    ws_summary['A12'] = "Total Net"; ws_summary['B12'] = mock_data['summary']['total_net']
    ws_summary.column_dimensions['A'].width = 24; ws_summary.column_dimensions['B'].width = 22

    ws_runs = wb.create_sheet("Payroll Runs", 1)
    _write_data_table(ws_runs, mock_data['payroll_runs'],
                      ["employee", "gross", "taxes", "benefits", "net", "pay_date"],
                      ["Employee", "Gross", "Taxes", "Benefits", "Net", "Pay Date"])


def _create_analysis_sheets(wb, case: Case, mock_data: dict):
    """Create sheets for analysis data."""
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "Analysis Case Summary"
    ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws_summary.merge_cells('A1:B1')
    
    ws_summary['A3'] = "Case ID"
    ws_summary['B3'] = mock_data['case_id']
    ws_summary['A4'] = "Case Title"
    ws_summary['B4'] = case.title
    ws_summary['A5'] = "Analysis Date"
    ws_summary['B5'] = mock_data['analysis_date']
    ws_summary['A6'] = "Generated On"
    ws_summary['B6'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    ws_summary['A8'] = "Summary Metrics"
    ws_summary['A8'].font = Font(bold=True, size=12)
    
    ws_summary['A9'] = "Total Issues Found"
    ws_summary['B9'] = mock_data['summary']['total_issues']
    ws_summary['A10'] = "Critical Issues"
    ws_summary['B10'] = mock_data['summary']['critical_issues']
    ws_summary['A11'] = "Affected Segments"
    ws_summary['B11'] = mock_data['summary']['affected_segments']
    ws_summary['A12'] = "Estimated Revenue Impact"
    ws_summary['B12'] = mock_data['summary']['estimated_impact']
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # Metrics sheet
    ws_metrics = wb.create_sheet("Metrics", 1)
    _write_data_table(ws_metrics, mock_data['metrics'],
                      ["metric", "current_value", "benchmark_value", "variance", "variance_percent", "trend"],
                      ["Metric", "Current Value", "Benchmark", "Variance", "Variance %", "Trend"])
    
    # Trends sheet
    ws_trends = wb.create_sheet("Trends", 2)
    _write_data_table(ws_trends, mock_data['trends'],
                      ["month", "value", "change"],
                      ["Month", "Value", "Change"])
    
    # Segments sheet
    ws_segments = wb.create_sheet("Segments", 3)
    _write_data_table(ws_segments, mock_data['segments'],
                      ["segment", "count", "revenue", "churn_rate", "satisfaction_score"],
                      ["Segment", "Count", "Revenue", "Churn Rate %", "Satisfaction Score"])
    
    # Anomalies sheet
    ws_anomalies = wb.create_sheet("Anomalies", 4)
    _write_data_table(ws_anomalies, mock_data['anomalies'],
                      ["type", "severity", "detected_date", "description", "impact"],
                      ["Type", "Severity", "Detected Date", "Description", "Impact"])


def _write_data_table(worksheet, data: list, columns: list, headers: list):
    """Write data to worksheet with formatting."""
    
    # Write headers
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Format numbers
            if isinstance(value, (int, float)) and col_name in ["amount", "revenue", "balance", "variance", "churn_rate", "satisfaction_score", "new_mrr", "churn_mrr", "expansion_mrr", "total_mrr", "projected_mrr", "gross", "taxes", "benefits", "net"]:
                if col_name in ["amount", "revenue", "balance", "new_mrr", "churn_mrr", "expansion_mrr", "total_mrr", "projected_mrr", "gross", "taxes", "benefits", "net"]:
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '0.00'
    
    # Auto-adjust column widths
    for col_idx, col_name in enumerate(columns, 1):
        max_length = max(len(str(d.get(col_name, ""))) for d in data) if data else 0
        max_length = max(max_length, len(headers[col_idx - 1]))
        worksheet.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 30)
